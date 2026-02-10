#!/usr/bin/env python3
"""
ФАЙЛ ДАННЫХ ДЛЯ FAQ
Версия 2.2 — Без pandas, с улучшенной обработкой ошибок
"""
import csv
import logging
import os
from typing import Dict, List
from openpyxl import load_workbook
from config import config

logger = logging.getLogger(__name__)

def get_faq_data() -> List[Dict]:
    """Загружает FAQ из CSV и Excel файлов без использования pandas"""
    faq_list = []
    
    try:
        # Получаем пути из конфигурации
        faq_file = config.get_faq_file()  # Путь типа data/faq.csv
        content_file = config.get_content_file()  # Путь типа data/контент.xlsx
        
        logger.info(f"📂 Загрузка данных: FAQ={faq_file}, Контент={content_file}")
        
        # 1. Загружаем основной FAQ из CSV (обязательный файл)
        csv_faq = _load_faq_from_csv(faq_file)
        if csv_faq:
            faq_list.extend(csv_faq)
        else:
            logger.warning("⚠️ CSV файл FAQ не загружен или пуст")
        
        # 2. Загружаем дополнительный контент из Excel (опционально)
        if os.path.exists(content_file):
            excel_faq = _load_content_from_excel(content_file)
            if excel_faq:
                faq_list.extend(excel_faq)
        else:
            logger.info(f"ℹ️ Файл контента не найден: {content_file} (не критично)")
        
        # Если ничего не загрузилось, используем демо-данные
        if not faq_list:
            logger.warning("⚠️ Не загружено ни одного вопроса, использую демо-данные")
            faq_list = _get_demo_faq()
        else:
            logger.info(f"✅ Загружено {len(faq_list)} вопросов (без pandas)")
        
    except Exception as e:
        logger.error(f"❌ Критическая ошибка загрузки данных: {e}", exc_info=True)
        faq_list = _get_demo_faq()
    
    return faq_list

def _load_faq_from_csv(file_path: str) -> List[Dict]:
    """Загрузка FAQ из CSV файла"""
    faq_items = []
    
    try:
        if not os.path.exists(file_path):
            logger.warning(f"⚠️ CSV файл не найден: {file_path}")
            return faq_items
        
        with open(file_path, 'r', encoding='utf-8-sig') as csv_file:
            reader = csv.DictReader(csv_file, delimiter=',')
            for row_num, row in enumerate(reader, start=1):
                try:
                    # Гибкое сопоставление названий столбцов (рус/англ)
                    faq_item = {
                        'question': str(row.get('question', row.get('Вопрос', ''))).strip(),
                        'answer': str(row.get('answer', row.get('Ответ', ''))).strip(),
                        'keywords': str(row.get('keywords', row.get('Ключевые слова', ''))).strip(),
                        'norm_keywords': str(row.get('norm_keywords', row.get('Норм ключевые', ''))).strip(),
                        'norm_question': str(row.get('norm_question', row.get('Норм вопрос', ''))).strip(),
                        'category': str(row.get('category', row.get('Категория', 'Общее'))).strip()
                    }
                    
                    # Проверяем обязательные поля
                    if faq_item['question'] and faq_item['answer']:
                        faq_items.append(faq_item)
                    else:
                        logger.debug(f"⚠️ Пропущена строка {row_num}: отсутствует вопрос или ответ")
                        
                except Exception as e:
                    logger.warning(f"⚠️ Ошибка обработки строки {row_num} в CSV: {e}")
                    continue
                    
        logger.info(f"📄 Из CSV загружено {len(faq_items)} вопросов")
        
    except Exception as e:
        logger.error(f"❌ Ошибка чтения CSV: {e}", exc_info=True)
    
    return faq_items

def _load_content_from_excel(file_path: str) -> List[Dict]:
    """Загрузка дополнительного контента из Excel файла с улучшенной обработкой ошибок"""
    faq_items = []
    
    try:
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)
        ws = wb.active
        
        # ✅ УЛУЧШЕНИЕ: Обработка пустого файла
        try:
            first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        except StopIteration:
            logger.warning(f"⚠️ Excel файл {file_path} пустой или повреждён")
            wb.close()
            return faq_items
        
        # Получаем заголовки из первой строки
        headers = []
        for cell_value in first_row:
            header = str(cell_value).strip() if cell_value is not None else ''
            if not header:
                header = f"Column_{len(headers)+1}"
            headers.append(header)
        
        # ✅ УЛУЧШЕНИЕ: Проверка наличия заголовков
        if not any(headers):
            logger.warning(f"⚠️ Excel файл {file_path} не содержит заголовков")
            wb.close()
            return faq_items
        
        logger.debug(f"📊 Заголовки Excel: {headers}")
        
        # Итерируемся по остальным строкам
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Пропускаем полностью пустые строки
                if not any(cell is not None for cell in row):
                    continue
                
                row_data = {}
                for col_idx, header in enumerate(headers):
                    value = row[col_idx] if col_idx < len(row) else None
                    row_data[header] = str(value).strip() if value is not None else ""
                
                # Преобразуем структуру Excel в структуру FAQ
                faq_item = {
                    'question': row_data.get('Вопрос', row_data.get('question', '')),
                    'answer': row_data.get('Ответ', row_data.get('answer', '')),
                    'keywords': row_data.get('Ключевые слова', row_data.get('keywords', '')),
                    'norm_keywords': row_data.get('Норм ключевые', row_data.get('norm_keywords', '')),
                    'norm_question': row_data.get('Норм вопрос', row_data.get('norm_question', '')),
                    'category': row_data.get('Категория', row_data.get('category', 'Дополнительно'))
                }
                
                # Добавляем только если есть вопрос и ответ
                if faq_item['question'] and faq_item['answer']:
                    faq_items.append(faq_item)
                    
            except Exception as e:
                logger.warning(f"⚠️ Ошибка обработки строки {row_idx} в Excel: {e}")
                continue
        
        wb.close()
        
        if faq_items:
            logger.info(f"📊 Из Excel загружено {len(faq_items)} вопросов")
        
    except FileNotFoundError:
        logger.warning(f"⚠️ Excel файл не найден: {file_path}")
    except Exception as e:
        logger.error(f"❌ Ошибка чтения Excel {file_path}: {e}")
    
    return faq_items

def _get_demo_faq() -> List[Dict]:
    """Демо-данные на случай отсутствия файлов"""
    return [
        {
            'question': 'Как оформить отпуск?',
            'answer': 'Напишите заявление руководителю и передайте в отдел кадров.',
            'keywords': 'отпуск, оформление, заявление',
            'norm_keywords': 'отпуск оформление заявление',
            'norm_question': 'как оформить отпуск',
            'category': 'Кадры'
        },
        {
            'question': 'Когда выдают зарплату?',
            'answer': 'Зарплата выплачивается 5 и 20 числа каждого месяца.',
            'keywords': 'зарплата, выплата, дата',
            'norm_keywords': 'зарплата выплата дата',
            'norm_question': 'когда выдают зарплату',
            'category': 'Финансы'
        },
        {
            'question': 'Как получить справку 2-НДФЛ?',
            'answer': 'Закажите справку через внутренний портал или обратитесь в бухгалтерию.',
            'keywords': 'справка, 2-ндфл, документ',
            'norm_keywords': 'справка 2 ндфл документ',
            'norm_question': 'как получить справку 2 ндфл',
            'category': 'Документы'
        }
    ]

if __name__ == "__main__":
    # Тестирование загрузки
    import json
    logging.basicConfig(level=logging.INFO)
    data = get_faq_data()
    print(f"Загружено {len(data)} вопросов")
    if data:
        print(json.dumps(data[:3], indent=2, ensure_ascii=False))
