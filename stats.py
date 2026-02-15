# stats.py
"""
Модуль статистики для HR-бота Мечел
Версия 2.5 – финальная с улучшенными генераторами отчётов и очисткой кэша
"""
import asyncio
import io
import logging
from collections import defaultdict
from datetime import datetime, timedelta
from typing import Dict, List, Any, Tuple, Optional, Set

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from database import (
    log_daily_stat,
    add_response_time,
    log_error,
    save_rating as db_save_rating,
    get_recent_response_times,
    get_daily_stats_for_last_days,
    get_all_feedback,  # для загрузки отзывов
)

logger = logging.getLogger(__name__)

class BotStatistics:
    """
    Класс для сбора статистики с агрегацией в памяти и периодической записью в БД.
    Буфер ограничен 7 днями, сброс каждые 60 секунд.
    """

    def __init__(self, flush_interval: int = 60, max_buffer_days: int = 7):
        self.start_time = datetime.now()
        self.flush_interval = flush_interval
        self.max_buffer_days = max_buffer_days

        # Буферы для накопления статистики (in-memory)
        self._daily_buffer = defaultdict(lambda: {
            'messages': 0,
            'commands': 0,
            'searches': 0,
            'feedback': 0,
            'ratings_helpful': 0,
            'ratings_unhelpful': 0,
        })
        self._users_buffer = defaultdict(set)  # дата -> set user_id (для оперативного доступа)
        self._users_count_buffer = defaultdict(int)  # дата -> кол-во уникальных пользователей (из БД)
        self._response_times_cache = []  # последние 100 значений (для быстрого доступа)

        # Дополнительный буфер для точного подсчёта активных за 24ч
        self._user_last_active = {}  # user_id -> datetime последней активности

        # Загружаем последние 7 дней из БД для инициализации буфера
        asyncio.create_task(self._load_recent_stats())

        # Задача для периодического сброса
        self._flush_task: Optional[asyncio.Task] = None
        asyncio.create_task(self._start_flush_loop())

    async def _load_recent_stats(self):
        """Загружает статистику за последние 7 дней из БД."""
        try:
            stats = await get_daily_stats_for_last_days(self.max_buffer_days)
            for date, data in stats.items():
                self._daily_buffer[date]['messages'] = data['messages']
                self._daily_buffer[date]['commands'] = data['commands']
                self._daily_buffer[date]['searches'] = data['searches']
                self._daily_buffer[date]['feedback'] = data['feedback']
                self._daily_buffer[date]['ratings_helpful'] = data['ratings']['helpful']
                self._daily_buffer[date]['ratings_unhelpful'] = data['ratings']['unhelpful']
                self._users_count_buffer[date] = data['users_count']
            logger.info(f"✅ Загружена статистика за {len(stats)} дней из БД")
        except Exception as e:
            logger.error(f"❌ Ошибка загрузки статистики из БД: {e}")

    async def _start_flush_loop(self):
        """Запускает цикл периодического сброса данных в БД."""
        while True:
            await asyncio.sleep(self.flush_interval)
            await self.flush()

    async def flush(self):
        """Сбрасывает накопленные данные в БД и очищает старые дни."""
        logger.debug("Сброс статистики в БД...")
        for date, counts in list(self._daily_buffer.items()):
            for field, value in counts.items():
                if value > 0:
                    await log_daily_stat(date, field, value)
            counts.clear()

        for date, users in list(self._users_buffer.items()):
            if users:
                self._users_count_buffer[date] = len(users)
                await log_daily_stat(date, 'users_count', len(users))
            users.clear()

        cutoff = (datetime.now() - timedelta(days=self.max_buffer_days)).strftime("%Y-%m-%d")
        for date in list(self._daily_buffer.keys()):
            if date < cutoff:
                del self._daily_buffer[date]
        for date in list(self._users_buffer.keys()):
            if date < cutoff:
                del self._users_buffer[date]
        for date in list(self._users_count_buffer.keys()):
            if date < cutoff:
                del self._users_count_buffer[date]

        # Очистка старых записей из _user_last_active (старше max_buffer_days)
        cutoff_7d = datetime.now() - timedelta(days=self.max_buffer_days)
        old_keys = [uid for uid, last_active in self._user_last_active.items()
                    if last_active < cutoff_7d]
        for uid in old_keys:
            del self._user_last_active[uid]
        logger.debug(f"Очищено {len(old_keys)} старых записей из _user_last_active")

        logger.debug("Сброс статистики завершён.")

    # --- Методы логирования ---
    async def log_message(self, user_id: int, username: str, msg_type: str, text: str = ""):
        now = datetime.now()
        date_key = now.strftime("%Y-%m-%d")

        # Обновляем время последней активности
        self._user_last_active[user_id] = now

        if msg_type == 'command':
            self._daily_buffer[date_key]['commands'] += 1
        elif msg_type == 'message':
            self._daily_buffer[date_key]['messages'] += 1
        elif msg_type == 'search':
            self._daily_buffer[date_key]['searches'] += 1
        elif msg_type == 'feedback':
            self._daily_buffer[date_key]['feedback'] += 1
            from database import save_feedback
            await save_feedback(user_id, username, text)
        elif msg_type == 'rating_helpful':
            self._daily_buffer[date_key]['ratings_helpful'] += 1
        elif msg_type == 'rating_unhelpful':
            self._daily_buffer[date_key]['ratings_unhelpful'] += 1

        self._users_buffer[date_key].add(user_id)

    def track_response_time(self, response_time: float):
        """Записывает время ответа в кэш и в БД."""
        self._response_times_cache.append(response_time)
        if len(self._response_times_cache) > 100:
            self._response_times_cache.pop(0)
        asyncio.create_task(add_response_time(response_time))

    def get_avg_response_time(self) -> float:
        if not self._response_times_cache:
            return 0.0
        return sum(self._response_times_cache) / len(self._response_times_cache)

    def get_response_time_status(self) -> Tuple[str, str]:
        avg = self.get_avg_response_time()
        if avg < 1.0:
            return "Хорошо", "green"
        elif avg < 3.0:
            return "Нормально", "yellow"
        else:
            return "Медленно", "red"

    def log_error(self, error_type: str, error_msg: str, user_id: int = None):
        asyncio.create_task(log_error(error_type, error_msg, user_id))

    def record_rating(self, faq_id: int, is_helpful: bool):
        date_key = datetime.now().strftime("%Y-%m-%d")
        self._daily_buffer[date_key]['ratings_helpful' if is_helpful else 'ratings_unhelpful'] += 1
        asyncio.create_task(db_save_rating(faq_id, 0, is_helpful))

    async def get_rating_stats(self) -> Dict[str, Any]:
        from database import get_rating_stats as db_stats
        return await db_stats()

    def get_summary_stats(self, period: str = 'all', cache_size: int = 0) -> Dict[str, Any]:
        now = datetime.now()
        if period == 'all':
            total_users = sum(self._users_count_buffer.values())
            total_messages = sum(d['messages'] for d in self._daily_buffer.values())
            total_commands = sum(d['commands'] for d in self._daily_buffer.values())
            total_searches = sum(d['searches'] for d in self._daily_buffer.values())
            total_feedback = sum(d['feedback'] for d in self._daily_buffer.values())
            total_ratings_helpful = sum(d['ratings_helpful'] for d in self._daily_buffer.values())
            total_ratings_unhelpful = sum(d['ratings_unhelpful'] for d in self._daily_buffer.values())
            all_response_times = self._response_times_cache
        else:
            delta_map = {
                'day': timedelta(days=1),
                'week': timedelta(days=7),
                'month': timedelta(days=30),
                'quarter': timedelta(days=90),
                'halfyear': timedelta(days=180),
                'year': timedelta(days=365)
            }
            delta = delta_map.get(period, timedelta(days=30))
            cutoff = (now - delta).date()
            total_users = 0
            total_messages = total_commands = total_searches = total_feedback = 0
            total_ratings_helpful = total_ratings_unhelpful = 0

            for date_str, users_cnt in self._users_count_buffer.items():
                try:
                    d = datetime.strptime(date_str, "%Y-%m-%d").date()
                    if d >= cutoff:
                        total_users += users_cnt
                except:
                    continue

            for date_str, counts in self._daily_buffer.items():
                try:
                    d = datetime.strptime(date_str, "%Y-%m-%d").date()
                    if d >= cutoff:
                        total_messages += counts['messages']
                        total_commands += counts['commands']
                        total_searches += counts['searches']
                        total_feedback += counts['feedback']
                        total_ratings_helpful += counts['ratings_helpful']
                        total_ratings_unhelpful += counts['ratings_unhelpful']
                except:
                    continue

            all_response_times = self._response_times_cache

        avg_response_time = sum(all_response_times) / len(all_response_times) if all_response_times else 0
        status, color = self.get_response_time_status()

        # Подсчёт активных за 24 часа
        cutoff_24h = now - timedelta(hours=24)
        active_24h = sum(1 for last_active in self._user_last_active.values() if last_active >= cutoff_24h)

        return {
            'period': period,
            'uptime': str(now - self.start_time),
            'start_time': self.start_time.strftime("%Y-%m-%d %H:%M:%S"),
            'total_users': total_users,
            'active_users_24h': active_24h,
            'total_messages': total_messages,
            'total_commands': total_commands,
            'total_searches': total_searches,
            'total_feedback': total_feedback,
            'total_ratings_helpful': total_ratings_helpful,
            'total_ratings_unhelpful': total_ratings_unhelpful,
            'total_ratings': total_ratings_helpful + total_ratings_unhelpful,
            'avg_response_time': avg_response_time,
            'response_time_status': status,
            'response_time_color': color,
            'cache_size': cache_size,
            'error_count': 0
        }

    def get_total_users(self) -> int:
        return sum(self._users_count_buffer.values())

    def get_weekly_stats_html(self) -> str:
        rows = []
        sorted_dates = sorted(self._daily_buffer.keys(), reverse=True)[:7]
        for date in sorted_dates:
            counts = self._daily_buffer[date]
            users = self._users_buffer[date]
            rows.append(f"""
                <tr>
                    <td>{date}</td>
                    <td>{len(users)}</td>
                    <td>{counts['messages']}</td>
                    <td>{counts['commands']}</td>
                    <td>{counts['searches']}</td>
                    <td>0.00с</td>
                    <td>{counts['ratings_helpful']}</td>
                    <td>{counts['ratings_unhelpful']}</td>
                </tr>
            """)
        return ''.join(rows)

    async def shutdown(self):
        """При завершении принудительно сбрасываем данные."""
        await self.flush()


# ---------- Генераторы отчётов (синхронные) ----------
def generate_feedback_report(bot_stats: BotStatistics) -> io.BytesIO:
    """
    Генерирует Excel-файл с отзывами.
    ВНИМАНИЕ: функция синхронная, но для получения данных из БД требуется асинхронный вызов.
    В текущей реализации возвращается заглушка. Рекомендуется переделать на асинхронную версию.
    """
    output = io.BytesIO()
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Отзывы и предложения"
        headers = ["Дата", "User ID", "Имя пользователя", "Текст"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = h
            cell.font = Font(bold=True)

        # Попытка загрузить отзывы из БД (асинхронно) – не рекомендуется в синхронной функции.
        # Для работоспособности оставляем заглушку.
        ws.cell(row=2, column=1, value="Для загрузки отзывов используйте асинхронную версию или веб-интерфейс")

        wb.save(output)
        output.seek(0)
    except Exception as e:
        logger.error(f"Ошибка генерации отчёта по отзывам: {e}")
        # Возвращаем пустой файл с ошибкой
        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Ошибка"
        ws['A1'] = f"Ошибка: {e}"
        wb.save(output)
        output.seek(0)
    return output


def generate_excel_report(bot_stats: BotStatistics, subscribers: List[int], search_engine=None) -> io.BytesIO:
    """
    Полный экспорт в Excel.
    Возвращает BytesIO с готовым файлом.
    """
    output = io.BytesIO()
    try:
        wb = Workbook()
        stats = bot_stats.get_summary_stats() if bot_stats else {}

        # Лист 1: Общая статистика
        ws1 = wb.active
        ws1.title = "Общая статистика"
        ws1['A1'] = "Статистика HR-бота Мечел"
        ws1['A1'].font = Font(bold=True, size=14)
        ws1.merge_cells('A1:D1')
        ws1['A3'] = "Показатель"; ws1['B3'] = "Значение"
        for cell in ['A3','B3']: ws1[cell].font = Font(bold=True)
        rows = [
            ("Дата экспорта", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Время работы", stats.get('uptime', 'N/A')),
            ("Запущен", stats.get('start_time', 'N/A')),
            ("Всего пользователей", stats.get('total_users', 0)),
            ("Активные (24ч)", stats.get('active_users_24h', 0)),
            ("Всего сообщений", stats.get('total_messages', 0)),
            ("Всего команд", stats.get('total_commands', 0)),
            ("Всего поисков", stats.get('total_searches', 0)),
            ("Всего отзывов/предложений", stats.get('total_feedback', 0)),
            ("Всего оценок", stats.get('total_ratings', 0)),
            ("Полезных ответов", stats.get('total_ratings_helpful', 0)),
            ("Бесполезных ответов", stats.get('total_ratings_unhelpful', 0)),
            ("Удовлетворённость", f"{stats.get('total_ratings_helpful', 0) / max(stats.get('total_ratings', 1), 1) * 100:.1f}%"),
            ("Ср. время ответа", f"{stats.get('avg_response_time', 0):.2f} сек"),
            ("Статус времени", stats.get('response_time_status', 'N/A')),
            ("Размер кэша", stats.get('cache_size', 0)),
            ("Количество ошибок", stats.get('error_count', 0)),
            ("Подписчиков", len(subscribers))
        ]
        for i, (k, v) in enumerate(rows, 4):
            ws1[f'A{i}'] = k; ws1[f'B{i}'] = v

        # Лист 2: Время ответа (последние 100)
        ws2 = wb.create_sheet("Время ответа")
        ws2['A1'] = "История времени ответа (последние 100)"
        ws2['A1'].font = Font(bold=True, size=14)
        ws2.merge_cells('A1:C1')
        ws2['A3'] = "Время"; ws2['B3'] = "Ответ (сек)"; ws2['C3'] = "Статус"
        for c in ['A3','B3','C3']: ws2[c].font = Font(bold=True)
        if bot_stats:
            for i, rt in enumerate(bot_stats._response_times_cache, 4):
                ws2[f'A{i}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")  # точной метки нет
                ws2[f'B{i}'] = rt
                t = rt
                ws2[f'C{i}'] = "Хорошо" if t < 1 else "Нормально" if t < 3 else "Медленно"

        # Лист 3: База знаний FAQ
        ws3 = wb.create_sheet("FAQ База")
        ws3['A1'] = "База знаний FAQ"
        ws3['A1'].font = Font(bold=True, size=14)
        ws3.merge_cells('A1:E1')
        headers_faq = ["ID", "Категория", "Вопрос", "Ответ", "Ключевые слова"]
        for col, h in enumerate(headers_faq, 1):
            cell = ws3.cell(row=3, column=col); cell.value = h; cell.font = Font(bold=True)

        if search_engine and hasattr(search_engine, 'faq_data') and search_engine.faq_data:
            row = 4
            for item in search_engine.faq_data:
                item_id = item.get('id', '')
                cat = item.get('category', 'Без категории')
                q = item.get('question', '')
                a = item.get('answer', '')
                kw = item.get('keywords', '')
                ws3.cell(row=row, column=1, value=item_id)
                ws3.cell(row=row, column=2, value=cat)
                ws3.cell(row=row, column=3, value=q)
                ws3.cell(row=row, column=4, value=a)
                ws3.cell(row=row, column=5, value=kw)
                row += 1
        else:
            ws3.cell(row=4, column=1, value="Поисковый движок недоступен или база знаний пуста")

        # Лист 4: Пользователи (активные за последние 7 дней)
        ws4 = wb.create_sheet("Пользователи")
        ws4['A1'] = "Активные пользователи (последние 7 дней)"
        ws4['A1'].font = Font(bold=True, size=14)
        ws4.merge_cells('A1:D1')
        headers_users = ["User ID", "Последняя активность", "Подписан на рассылку"]
        for col, h in enumerate(headers_users, 1):
            cell = ws4.cell(row=3, column=col); cell.value = h; cell.font = Font(bold=True)

        if bot_stats:
            subs_set = set(subscribers)
            # Берём всех пользователей из _user_last_active (они уже не старше 7 дней из-за очистки)
            row = 4
            for uid, last_active in sorted(bot_stats._user_last_active.items(), key=lambda x: x[1], reverse=True):
                ws4.cell(row=row, column=1, value=uid)
                ws4.cell(row=row, column=2, value=last_active.strftime("%Y-%m-%d %H:%M:%S") if last_active else '')
                ws4.cell(row=row, column=3, value="Да" if uid in subs_set else "Нет")
                row += 1
                if row > 10000:  # Защита от слишком больших файлов
                    ws4.cell(row=row, column=1, value="... (слишком много пользователей, показаны первые 10000)")
                    break

        # Лист 5: Оценки FAQ (заглушка)
        ws5 = wb.create_sheet("Оценки FAQ")
        ws5['A1'] = "Статистика оценок по вопросам"
        ws5['A1'].font = Font(bold=True, size=14)
        ws5.merge_cells('A1:D1')
        headers_ratings = ["ID вопроса", "Вопрос", "👍 Помог", "👎 Нет", "Всего оценок"]
        for col, h in enumerate(headers_ratings, 1):
            cell = ws5.cell(row=3, column=col); cell.value = h; cell.font = Font(bold=True)
        ws5.cell(row=4, column=1, value="Для получения оценок используйте асинхронную версию")

        # Автоподбор ширины столбцов
        for ws in [ws1, ws2, ws3, ws4, ws5]:
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if cell.value and len(str(cell.value)) > max_len:
                            max_len = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = min(max_len + 2, 70)

        wb.save(output)
        output.seek(0)
    except Exception as e:
        logger.error(f"Ошибка генерации Excel-отчёта: {e}", exc_info=True)
        # Возвращаем пустой файл с информацией об ошибке
        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Ошибка"
        ws['A1'] = f"Ошибка при формировании отчёта: {e}"
        wb.save(output)
        output.seek(0)
    return output
